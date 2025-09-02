import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook

st.set_page_config(page_title="Application Form", layout="wide")
st.title("📋 Grant Application Portal")

# --- Session State Setup ---
if "logframe" not in st.session_state:
    st.session_state["logframe"] = []
if "workplan" not in st.session_state:
    st.session_state["workplan"] = []
if "budget" not in st.session_state:
    st.session_state["budget"] = []

# --- File Upload / Resume Editing ---
st.sidebar.header("🔁 Resume Previous Submission")
uploaded_file = st.sidebar.file_uploader("Upload previously saved Excel", type="xlsx")
if uploaded_file:
    wb = load_workbook(uploaded_file)
    st.session_state["logframe"] = pd.DataFrame(wb["Logframe"].values)[1:].values.tolist()
    st.session_state["workplan"] = pd.DataFrame(wb["Workplan"].values)[1:].values.tolist()
    st.session_state["budget"] = pd.DataFrame(wb["Budget"].values)[1:].values.tolist()
    st.sidebar.success("✅ Data loaded from Excel!")

# --- Step 1: Logframe ---
st.header("Step 1: Logframe")
with st.form("logframe_form", clear_on_submit=True):
    col1, col2, col3 = st.columns(3)
    goal = col1.text_input("Goal")
    outcome = col2.text_input("Outcome")
    output = col3.text_input("Output")
    indicator = st.text_input("Indicator")
    target = st.text_input("Target")
    means = st.text_input("Means of Verification")
    submitted_log = st.form_submit_button("➕ Add to Logframe")
    if submitted_log:
        st.session_state["logframe"].append([goal, outcome, output, indicator, target, means])
        st.success("Added to logframe")

if st.session_state["logframe"]:
    st.subheader("Current Logframe Entries")
    st.table(st.session_state["logframe"])

# --- Step 2: Workplan ---
st.header("Step 2: Workplan")
with st.form("workplan_form", clear_on_submit=True):
    activity = st.text_input("Activity")
    linked_output = st.selectbox("Linked Output", [row[2] for row in st.session_state["logframe"]] or ["None"])
    start = st.date_input("Start Date")
    end = st.date_input("End Date")
    person = st.text_input("Responsible Person")
    milestone = st.text_input("Milestone")
    submitted_workplan = st.form_submit_button("➕ Add to Workplan")
    if submitted_workplan:
        st.session_state["workplan"].append([activity, linked_output, str(start), str(end), person, milestone])
        st.success("Added to workplan")

if st.session_state["workplan"]:
    st.subheader("Current Workplan Entries")
    st.table(st.session_state["workplan"])

# --- Step 3: Budget ---
st.header("Step 3: Budget")
with st.form("budget_form", clear_on_submit=True):
    line_item = st.text_input("Line Item")
    linked_activity = st.selectbox("Linked Activity", [row[0] for row in st.session_state["workplan"]] or ["None"])
    category = st.selectbox("Category", ["Personnel", "Supplies", "Travel", "Other"])
    unit = st.text_input("Unit")
    quantity = st.number_input("Quantity", min_value=0.0)
    unit_cost = st.number_input("Unit Cost", min_value=0.0)
    total_cost = quantity * unit_cost
    submitted_budget = st.form_submit_button("➕ Add to Budget")
    if submitted_budget:
        st.session_state["budget"].append([line_item, linked_activity, category, unit, quantity, unit_cost, total_cost])
        st.success("Added to budget")

if st.session_state["budget"]:
    st.subheader("Current Budget Entries")
    st.table(st.session_state["budget"])

# --- Download Final Excel ---
def generate_excel():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Logframe"
    ws1.append(["Goal", "Outcome", "Output", "Indicator", "Target", "Means of Verification"])
    for row in st.session_state["logframe"]:
        ws1.append(row)

    ws2 = wb.create_sheet("Workplan")
    ws2.append(["Activity", "Linked Output", "Start Date", "End Date", "Responsible Person", "Milestone"])
    for row in st.session_state["workplan"]:
        ws2.append(row)

    ws3 = wb.create_sheet("Budget")
    ws3.append(["Line Item", "Linked Activity", "Category", "Unit", "Quantity", "Unit Cost", "Total Cost"])
    for row in st.session_state["budget"]:
        ws3.append(row)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

st.header("📥 Export Your Application")
if st.button("📤 Generate Excel File"):
    excel_file = generate_excel()
    st.download_button(
        label="Download Completed Template",
        data=excel_file,
        file_name="My_Application.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
